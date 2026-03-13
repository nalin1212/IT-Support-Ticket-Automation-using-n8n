"""
=============================================================
  Assessment 1 — IT Support Ticket Automation System
  University IT Department
=============================================================
  Features:
    - CSV file trigger / manual ticket input
    - Email & priority validation
    - Text normalization
    - Deduplication (same email + issue within 24h)
    - Unique Ticket ID generation
    - Team routing (wifi/login/software/hardware/other)
    - SLA deadline calculation (High=4h, Medium=24h, Low=72h)
    - Error handling for all edge cases
    - Outputs: processed.csv, rejected.csv, summary_report.xlsx
=============================================================
"""

import csv
import re
import uuid
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict

try:
    import pandas as pd
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ─────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────
INPUT_FILE       = "tickets_input.csv"
PROCESSED_FILE   = "processed_tickets.csv"
REJECTED_FILE    = "rejected_tickets.csv"
REPORT_FILE      = "summary_report.xlsx"
REPORT_CSV_FILE  = "summary_report.csv"

VALID_ISSUE_TYPES = {"wifi", "login", "software", "hardware", "other"}
VALID_PRIORITIES  = {"High", "Medium", "Low"}

ROUTING_MAP = {
    "wifi":     "Network",
    "login":    "IT Support",
    "software": "Applications",
    "hardware": "Infrastructure",
    "other":    "General",
}

SLA_HOURS = {"High": 4, "Medium": 24, "Low": 72}

# ─────────────────────────────────────────
#  UTILITIES
# ─────────────────────────────────────────
def generate_ticket_id():
    """Generate a unique ticket ID: TKT-YYYYMMDD-XXXXXXXX"""
    date_part = datetime.now().strftime("%Y%m%d")
    unique_part = uuid.uuid4().hex[:8].upper()
    return f"TKT-{date_part}-{unique_part}"

def validate_email(email: str) -> bool:
    """Validate email using RFC-compliant regex."""
    pattern = r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email.strip()))

def normalize_text(text: str) -> str:
    """Strip whitespace and title-case a name field."""
    return " ".join(text.strip().split())

def normalize_name(name: str) -> str:
    """Title-case name."""
    return normalize_text(name).title()

def normalize_issue_type(issue: str) -> str:
    """Lowercase and strip issue type."""
    return normalize_text(issue).lower()

def normalize_priority(priority: str) -> str:
    """Title-case priority (high->High)."""
    return normalize_text(priority).title()

def calculate_sla(priority: str, timestamp: datetime) -> str:
    """Return SLA deadline as formatted string."""
    hours = SLA_HOURS.get(priority, 24)
    deadline = timestamp + timedelta(hours=hours)
    return deadline.strftime("%Y-%m-%d %H:%M:%S")

def dedup_key(email: str, issue_type: str) -> str:
    return f"{email.lower()}|{issue_type.lower()}"

def parse_timestamp(ts_str: str) -> datetime:
    """Try multiple formats to parse timestamp."""
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(ts_str.strip(), fmt)
        except ValueError:
            continue
    return datetime.now()  # fallback


# ─────────────────────────────────────────
#  CORE PROCESSOR
# ─────────────────────────────────────────
class TicketProcessor:
    def __init__(self):
        self.processed = []
        self.rejected  = []
        self.seen_tickets = {}   # dedup_key -> last submission datetime
        self.stats = defaultdict(int)

    # ── VALIDATE ──────────────────────────
    def validate(self, row: dict, row_num: int) -> tuple[bool, str]:
        """Run all validations. Returns (is_valid, reason_if_not)."""
        name       = row.get("Name", "").strip()
        email      = row.get("Email", "").strip()
        issue_type = row.get("Issue Type", "").strip()
        priority   = row.get("Priority", "").strip()
        desc       = row.get("Description", "").strip()

        if not name:
            return False, "Name is missing"

        if not email:
            return False, "Email is missing"
        if not validate_email(email):
            return False, f"Invalid email format: '{email}'"

        norm_issue = normalize_issue_type(issue_type)
        if not norm_issue:
            return False, "Issue type is missing"
        if norm_issue not in VALID_ISSUE_TYPES:
            return False, f"Unknown issue type: '{issue_type}' (allowed: {', '.join(VALID_ISSUE_TYPES)})"

        norm_priority = normalize_priority(priority)
        if not priority:
            return False, "Priority is missing"
        if norm_priority not in VALID_PRIORITIES:
            return False, f"Invalid priority: '{priority}' (allowed: High, Medium, Low)"

        if not desc:
            return False, "Description is missing"

        return True, ""

    # ── DEDUPLICATE ───────────────────────
    def is_duplicate(self, email: str, issue_type: str, timestamp: datetime) -> bool:
        key = dedup_key(email, issue_type)
        if key in self.seen_tickets:
            last_time = self.seen_tickets[key]
            diff_hours = (timestamp - last_time).total_seconds() / 3600
            if diff_hours < 24:
                return True
        self.seen_tickets[key] = timestamp
        return False

    # ── PROCESS ONE TICKET ─────────────────
    def process_ticket(self, row: dict, row_num: int):
        self.stats["total"] += 1

        # Step 1: Parse timestamp
        ts_raw = row.get("Timestamp", "").strip()
        timestamp = parse_timestamp(ts_raw) if ts_raw else datetime.now()

        # Step 2: Validate
        is_valid, reason = self.validate(row, row_num)
        if not is_valid:
            self._reject(row, reason, timestamp)
            return

        # Step 3: Normalize fields
        name       = normalize_name(row["Name"])
        email      = row["Email"].strip().lower()
        issue_type = normalize_issue_type(row["Issue Type"])
        priority   = normalize_priority(row["Priority"])
        desc       = normalize_text(row["Description"])

        # Step 4: Deduplication
        if self.is_duplicate(email, issue_type, timestamp):
            self._reject(
                row,
                f"Duplicate ticket — same email+issue '{issue_type}' already submitted within 24 hours",
                timestamp
            )
            return

        # Step 5: Generate unique ID
        ticket_id = generate_ticket_id()

        # Step 6: Route to team
        team = ROUTING_MAP[issue_type]

        # Step 7: Calculate SLA
        sla_deadline = calculate_sla(priority, timestamp)

        # Step 8: Store processed ticket
        processed_ticket = {
            "Ticket ID":    ticket_id,
            "Name":         name,
            "Email":        email,
            "Issue Type":   issue_type,
            "Priority":     priority,
            "Description":  desc,
            "Routed To":    team,
            "SLA Deadline": sla_deadline,
            "Submitted At": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "Status":       "Processed",
        }
        self.processed.append(processed_ticket)
        self.stats["processed"] += 1
        self.stats[f"team_{team}"] += 1
        self.stats[f"priority_{priority}"] += 1

        print(f"  ✓ [{ticket_id}] {name} → {team} (SLA: {sla_deadline})")

    def _reject(self, row: dict, reason: str, timestamp: datetime):
        """Store a rejected ticket with its reason."""
        rejected_ticket = {
            "Ticket ID":  generate_ticket_id().replace("TKT", "REJ"),
            "Name":       row.get("Name", "").strip() or "—",
            "Email":      row.get("Email", "").strip() or "—",
            "Issue Type": row.get("Issue Type", "").strip() or "—",
            "Priority":   row.get("Priority", "").strip() or "—",
            "Description":row.get("Description","").strip() or "—",
            "Reason":     reason,
            "Submitted At": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            "Status":     "Rejected",
        }
        self.rejected.append(rejected_ticket)
        self.stats["rejected"] += 1
        name_display = row.get("Name","").strip() or "Unknown"
        print(f"  ✗ [{rejected_ticket['Ticket ID']}] {name_display} → REJECTED: {reason}")

    # ── PROCESS CSV FILE ──────────────────
    def process_csv(self, filepath: str):
        """Read and process all tickets from a CSV file."""
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Input file not found: {filepath}")

        print(f"\n{'='*60}")
        print(f"  TICKET AUTOMATION SYSTEM — Processing: {filepath}")
        print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}\n")

        try:
            with open(filepath, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        except Exception as e:
            raise RuntimeError(f"Failed to read CSV: {e}")

        if not rows:
            raise ValueError("CSV file is empty or has no data rows.")

        print(f"  Loaded {len(rows)} ticket(s) from file.\n")

        for i, row in enumerate(rows, start=1):
            print(f"  [Row {i:02d}] Processing: {row.get('Name','?')} | {row.get('Email','?')} | {row.get('Issue Type','?')}")
            self.process_ticket(row, i)

        print(f"\n{'='*60}")
        print(f"  PROCESSING COMPLETE")
        print(f"  Total:     {self.stats['total']}")
        print(f"  Processed: {self.stats['processed']}")
        print(f"  Rejected:  {self.stats['rejected']}")
        print(f"{'='*60}\n")

    # ── SAVE OUTPUTS ──────────────────────
    def save_processed_csv(self, filepath: str):
        if not self.processed:
            print("  ⚠ No processed tickets to save.")
            return
        fieldnames = list(self.processed[0].keys())
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.processed)
        print(f"  ✓ Processed tickets → {filepath}  ({len(self.processed)} records)")

    def save_rejected_csv(self, filepath: str):
        if not self.rejected:
            print("  ✓ No rejected tickets.")
            return
        fieldnames = list(self.rejected[0].keys())
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.rejected)
        print(f"  ✓ Rejected tickets  → {filepath}  ({len(self.rejected)} records)")

    def save_summary_report(self, xlsx_path: str, csv_path: str):
        """Generate summary report in both Excel and CSV."""
        total     = self.stats["total"]
        processed = self.stats["processed"]
        rejected  = self.stats["rejected"]

        teams = {
            "Network":        self.stats.get("team_Network", 0),
            "IT Support":     self.stats.get("team_IT Support", 0),
            "Applications":   self.stats.get("team_Applications", 0),
            "Infrastructure": self.stats.get("team_Infrastructure", 0),
            "General":        self.stats.get("team_General", 0),
        }
        priorities = {
            "High":   self.stats.get("priority_High", 0),
            "Medium": self.stats.get("priority_Medium", 0),
            "Low":    self.stats.get("priority_Low", 0),
        }

        # ── Save CSV summary ──
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(["SUMMARY REPORT"])
            w.writerow(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            w.writerow([])
            w.writerow(["OVERVIEW"])
            w.writerow(["Total Tickets Received", total])
            w.writerow(["Processed", processed])
            w.writerow(["Rejected", rejected])
            w.writerow([])
            w.writerow(["TICKETS PER TEAM"])
            w.writerow(["Team", "Count"])
            for team, count in teams.items():
                w.writerow([team, count])
            w.writerow([])
            w.writerow(["TICKETS BY PRIORITY"])
            w.writerow(["Priority", "Count"])
            for p, count in priorities.items():
                w.writerow([p, count])
        print(f"  ✓ Summary CSV       → {csv_path}")

        # ── Save Excel ──
        if not EXCEL_AVAILABLE:
            print("  ⚠ pandas/openpyxl not available — skipping Excel report.")
            return

        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            # Sheet 1: Overview
            overview_df = pd.DataFrame([
                ["Total Tickets Received", total],
                ["Processed",              processed],
                ["Rejected",               rejected],
                ["Processing Rate",        f"{(processed/total*100):.1f}%" if total else "0%"],
                ["Rejection Rate",         f"{(rejected/total*100):.1f}%" if total else "0%"],
            ], columns=["Metric", "Value"])
            overview_df.to_excel(writer, sheet_name="Overview", index=False)

            # Sheet 2: Tickets Per Team
            teams_df = pd.DataFrame([
                {"Team": t, "Ticket Count": c, "Percentage": f"{(c/processed*100):.1f}%" if processed else "0%"}
                for t, c in teams.items()
            ])
            teams_df.to_excel(writer, sheet_name="Tickets Per Team", index=False)

            # Sheet 3: Priority Breakdown
            prio_df = pd.DataFrame([
                {"Priority": p, "Count": c, "SLA (hours)": SLA_HOURS.get(p,"—")}
                for p, c in priorities.items()
            ])
            prio_df.to_excel(writer, sheet_name="Priority Breakdown", index=False)

            # Sheet 4: Processed Tickets
            if self.processed:
                pd.DataFrame(self.processed).to_excel(writer, sheet_name="Processed Tickets", index=False)

            # Sheet 5: Rejected Tickets
            if self.rejected:
                pd.DataFrame(self.rejected).to_excel(writer, sheet_name="Rejected Tickets", index=False)

            # Style the sheets
            wb = writer.book
            from openpyxl.styles import PatternFill, Font, Alignment
            header_fill   = PatternFill("solid", fgColor="1F3864")
            header_font   = Font(color="FFFFFF", bold=True)
            alt_fill      = PatternFill("solid", fgColor="EBF3FB")

            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
                for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if i % 2 == 0:
                        for cell in row:
                            cell.fill = alt_fill

        print(f"  ✓ Summary Excel     → {xlsx_path}  (5 sheets)")


# ─────────────────────────────────────────
#  MAIN ENTRY POINT
# ─────────────────────────────────────────
def main():
    processor = TicketProcessor()

    # Step 1: Process input CSV
    processor.process_csv(INPUT_FILE)

    # Step 2: Save all outputs
    print("  Saving outputs...\n")
    processor.save_processed_csv(PROCESSED_FILE)
    processor.save_rejected_csv(REJECTED_FILE)
    processor.save_summary_report(REPORT_FILE, REPORT_CSV_FILE)

    print(f"\n{'='*60}")
    print("  ALL OUTPUTS SAVED SUCCESSFULLY")
    print(f"  → {PROCESSED_FILE}")
    print(f"  → {REJECTED_FILE}")
    print(f"  → {REPORT_FILE}")
    print(f"  → {REPORT_CSV_FILE}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
